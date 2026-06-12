"""
Deep inspection of geospatial files — extracts everything the agent creator needs.
"""
import csv as csv_mod
import json
import os
import zipfile


def inspect(filepath: str) -> dict:
    """Inspect any geo file and return a detailed metadata dict."""
    ext = os.path.splitext(filepath)[1].lower().lstrip(".")
    info = {
        "path": os.path.abspath(filepath),
        "name": os.path.basename(filepath),
        "type": ext,
        "size": os.path.getsize(filepath),
    }

    try:
        handler = {
            "kml": _inspect_kml,
            "kmz": _inspect_kmz,
            "tif": _inspect_tif,
            "tiff": _inspect_tif,
            "geojson": _inspect_geojson,
            "csv": _inspect_csv,
            "xlsx": _inspect_xlsx,
            "xls": _inspect_xlsx,
            "shp": _inspect_shp,
            "json": _inspect_json,
        }.get(ext)
        if handler:
            info.update(handler(filepath))
        else:
            info["note"] = f"No inspector for .{ext}"
    except Exception as e:
        info["inspect_error"] = str(e)

    return info


def _inspect_kml(path):
    from lxml import etree
    ns = "{http://www.opengis.net/kml/2.2}"
    tree = etree.parse(path, etree.XMLParser(recover=True, huge_tree=True))
    root = tree.getroot()

    features = list(root.iter(f"{ns}Placemark"))
    geom_types = {}
    for tag in ("Point", "LineString", "Polygon", "MultiGeometry", "LinearRing"):
        c = len(list(root.iter(f"{ns}{tag}")))
        if c:
            geom_types[tag] = c

    coords = []
    for el in root.iter(f"{ns}coordinates"):
        if el.text:
            for t in el.text.strip().split():
                p = t.split(",")
                if len(p) >= 2:
                    try:
                        coords.append((float(p[0]), float(p[1])))
                    except ValueError:
                        pass

    bounds = None
    if coords:
        lons, lats = zip(*coords)
        bounds = [min(lons), min(lats), max(lons), max(lats)]

    # Extract feature names
    names = []
    for pm in features[:20]:
        n = pm.find(f"{ns}name")
        if n is not None and n.text:
            names.append(n.text.strip())

    return {
        "file_format": "kml",
        "feature_count": len(features),
        "geometry_types": geom_types,
        "bounds": bounds,
        "vertex_count": len(coords),
        "feature_names": names,
        "coordinate_sample": coords[:10],
    }


def _inspect_kmz(path):
    meta = {"file_format": "kmz"}
    with zipfile.ZipFile(path, "r") as z:
        all_files = z.namelist()
        kml_files = [n for n in all_files if n.lower().endswith(".kml")]
        meta["contained_files"] = all_files
        meta["kml_files"] = kml_files

        if kml_files:
            import tempfile
            with tempfile.TemporaryDirectory() as tmp:
                z.extract(kml_files[0], tmp)
                kml_meta = _inspect_kml(os.path.join(tmp, kml_files[0]))
                meta.update(kml_meta)
                meta["file_format"] = "kmz"
    return meta


def _inspect_tif(path):
    import rasterio
    import numpy as np

    with rasterio.open(path) as ds:
        meta = {
            "file_format": "geotiff",
            "bands": ds.count,
            "width": ds.width,
            "height": ds.height,
            "crs": str(ds.crs),
            "crs_epsg": ds.crs.to_epsg() if ds.crs else None,
            "bounds": [round(b, 6) for b in ds.bounds],
            "resolution": [round(r, 6) for r in ds.res],
            "dtype": str(ds.dtypes[0]),
            "nodata": ds.nodata,
            "stats": {},
            "data_class": "unknown",
        }

        # Try to read a small sample for stats — skip if file is too large or GDAL errors
        try:
            h = min(ds.height, 256)
            w = min(ds.width, 256)
            data = ds.read(1, out_shape=(h, w))
            valid = data[data != ds.nodata] if ds.nodata is not None else data.ravel()

            if valid.size > 0:
                meta["stats"] = {
                    "min": float(np.nanmin(valid)),
                    "max": float(np.nanmax(valid)),
                    "mean": float(np.nanmean(valid)),
                    "std": float(np.nanstd(valid)),
                    "percentile_25": float(np.nanpercentile(valid, 25)),
                    "percentile_75": float(np.nanpercentile(valid, 75)),
                }

            stats = meta["stats"]
            if ds.count == 1 and ds.dtypes[0] in ("float32", "float64"):
                if 0 < stats.get("max", 0) < 9000:
                    meta["data_class"] = "elevation/DEM"
                elif stats.get("max", 0) - stats.get("min", 0) < 100:
                    meta["data_class"] = "slope/continuous"
            elif ds.count == 1 and "int" in str(ds.dtypes[0]):
                unique_approx = len(np.unique(data[:100, :100]))
                if unique_approx < 30:
                    meta["data_class"] = "classification"
            elif ds.count >= 3:
                meta["data_class"] = "imagery/RGB"
        except Exception:
            # File too large or GDAL error reading pixels — metadata only
            if ds.count >= 3:
                meta["data_class"] = "imagery/RGB"
            meta["stats_note"] = "skipped (file too large)"

        return meta


def _inspect_csv(path):
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv_mod.reader(f)
        headers = next(reader, [])
        rows = []
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= 200:
                break

    # Detect column types
    col_info = {}
    for ci, col in enumerate(headers):
        vals = [r[ci] for r in rows if ci < len(r) and r[ci].strip()]
        numeric = []
        for v in vals:
            try:
                numeric.append(float(v))
            except (ValueError, TypeError):
                pass

        is_numeric = len(numeric) > len(vals) * 0.7
        info = {"name": col, "is_numeric": is_numeric, "non_empty": len(vals)}
        if is_numeric and numeric:
            info["min"] = min(numeric)
            info["max"] = max(numeric)
            info["mean"] = sum(numeric) / len(numeric)
            # Check monotonic
            if len(numeric) > 2:
                diffs = [numeric[i+1] - numeric[i] for i in range(min(len(numeric)-1, 50))]
                info["monotonic_increasing"] = all(d >= 0 for d in diffs)
                if len(numeric) > 5:
                    intervals = [abs(d) for d in diffs if abs(d) > 0]
                    if intervals:
                        info["typical_interval"] = sum(intervals) / len(intervals)
        col_info[col] = info

    # Detect lat/lon
    lat_cols = [c for c in headers if c.lower() in ("lat", "latitude", "y", "lat_y", "lat_dd")]
    lon_cols = [c for c in headers if c.lower() in ("lon", "lng", "longitude", "x", "lon_x", "lon_dd")]

    sample_rows = []
    for row in rows[:20]:
        d = {}
        for ci, col in enumerate(headers):
            d[col] = row[ci] if ci < len(row) else ""
        sample_rows.append(d)

    return {
        "file_format": "csv",
        "rows": len(rows),
        "columns": headers,
        "column_info": col_info,
        "has_latlon": bool(lat_cols and lon_cols),
        "lat_columns": lat_cols,
        "lon_columns": lon_cols,
        "sample_rows": sample_rows,
    }


def _inspect_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {"file_format": "excel", "sheets": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
            else:
                rows_data.append(list(row))
            if i >= 50:
                break

        col_info = {}
        for ci, col in enumerate(headers):
            vals = [r[ci] for r in rows_data if ci < len(r) and r[ci] is not None]
            numeric = [v for v in vals if isinstance(v, (int, float))]
            info = {"name": col, "is_numeric": len(numeric) > len(vals) * 0.5}
            if numeric:
                info["min"] = min(numeric)
                info["max"] = max(numeric)
                info["mean"] = sum(numeric) / len(numeric)
            col_info[col] = info

        sample_rows = []
        for row in rows_data[:20]:
            d = {}
            for ci, col in enumerate(headers):
                d[col] = row[ci] if ci < len(row) else None
            sample_rows.append(d)

        result["sheets"][sheet_name] = {
            "rows": ws.max_row - 1 if ws.max_row else 0,
            "columns": headers,
            "column_info": col_info,
            "sample_rows": sample_rows,
        }

    wb.close()
    # Flatten if single sheet
    if len(result["sheets"]) == 1:
        sheet = list(result["sheets"].values())[0]
        result.update(sheet)
    return result


def _inspect_geojson(path):
    with open(path) as f:
        data = json.load(f)
    features = data.get("features", [])
    gtypes = {}
    for feat in features:
        gt = feat.get("geometry", {}).get("type", "unknown")
        gtypes[gt] = gtypes.get(gt, 0) + 1

    props = set()
    for feat in features[:50]:
        props.update(feat.get("properties", {}).keys())

    return {
        "file_format": "geojson",
        "feature_count": len(features),
        "geometry_types": gtypes,
        "properties": sorted(props),
    }


def _inspect_shp(path):
    try:
        import geopandas as gpd
        gdf = gpd.read_file(path)
        return {
            "file_format": "shapefile",
            "feature_count": len(gdf),
            "geometry_types": dict(gdf.geom_type.value_counts()),
            "columns": list(gdf.columns),
            "crs": str(gdf.crs) if gdf.crs else None,
            "bounds": list(gdf.total_bounds),
        }
    except Exception:
        return {"file_format": "shapefile"}


def _inspect_json(path):
    with open(path) as f:
        data = json.load(f)
    if isinstance(data, dict) and "features" in data:
        return _inspect_geojson(path)
    return {"file_format": "json", "keys": list(data.keys()) if isinstance(data, dict) else "array"}
