"""
Runner — execute a single pipeline step in the sandbox and collect results.
Also extracts metadata from output files for downstream agents (visualizer, etc).
"""
import json
import os

from geodb.transform.pipeline.step import Step
from geodb.transform.pipeline.sandbox import Sandbox


def run_step(step: Step, file_map: dict, timeout: int = None) -> dict:
    """
    Execute a step's code in a sandbox.

    Args:
        step: Step with .code populated
        file_map: { 'filename': '/path/to/actual/file' } for all inputs
        timeout: override execution timeout

    Returns:
        {
            success: bool,
            stdout: str,
            stderr: str,
            output_files: { filename: full_path },
            elapsed: float,
            error: str,
            output_metadata: dict,   # auto-extracted stats
        }
    """
    sandbox = Sandbox()

    try:
        # Copy inputs into sandbox
        needed = {}
        for inp in step.inputs:
            if inp in file_map:
                needed[inp] = file_map[inp]
            else:
                return {
                    "success": False, "stdout": "", "stderr": "",
                    "output_files": {}, "elapsed": 0.0,
                    "error": f"Missing input file: {inp}",
                    "output_metadata": {},
                    "sandbox": sandbox,
                }
        sandbox.setup_inputs(needed)

        # Execute
        result = sandbox.execute(step.code, timeout=timeout)

        # Build output file paths
        output_paths = sandbox.get_all_output_paths()

        # Extract metadata from outputs
        output_meta = {}
        for fname, fpath in output_paths.items():
            output_meta[fname] = _extract_metadata(fpath)

        return {
            "success": result["success"],
            "stdout": result["stdout"],
            "stderr": result["stderr"],
            "output_files": output_paths,
            "elapsed": result["elapsed"],
            "error": result["error"],
            "output_metadata": output_meta,
            "sandbox": sandbox,
        }

    except Exception as e:
        return {
            "success": False, "stdout": "", "stderr": "",
            "output_files": {}, "elapsed": 0.0,
            "error": str(e),
            "output_metadata": {},
            "sandbox": sandbox,
        }


def _extract_metadata(filepath: str) -> dict:
    """
    Quick metadata extraction from an output file.
    Used by the visualizer agent to understand what to render.
    """
    meta = {
        "path": filepath,
        "size_bytes": os.path.getsize(filepath),
        "extension": os.path.splitext(filepath)[1].lower(),
    }
    ext = meta["extension"]

    try:
        if ext == ".geojson":
            meta.update(_inspect_geojson(filepath))
        elif ext == ".csv":
            meta.update(_inspect_csv(filepath))
        elif ext in (".xlsx", ".xls"):
            meta.update(_inspect_excel(filepath))
        elif ext in (".tif", ".tiff"):
            meta.update(_inspect_tif(filepath))
        elif ext == ".json":
            meta["type"] = "json"
    except Exception as e:
        meta["inspect_error"] = str(e)

    return meta


def _inspect_geojson(path: str) -> dict:
    with open(path, "r") as f:
        data = json.load(f)

    features = data.get("features", [])
    geom_types = set()
    for feat in features:
        g = feat.get("geometry", {})
        if g:
            geom_types.add(g.get("type", "unknown"))

    # Compute bounds
    coords_flat = []
    for feat in features:
        _collect_coords(feat.get("geometry", {}), coords_flat)

    bounds = None
    if coords_flat:
        lons = [c[0] for c in coords_flat]
        lats = [c[1] for c in coords_flat]
        bounds = [min(lons), min(lats), max(lons), max(lats)]

    props = set()
    for feat in features:
        props.update(feat.get("properties", {}).keys())

    return {
        "type": "geojson",
        "feature_count": len(features),
        "geometry_types": list(geom_types),
        "bounds": bounds,
        "properties": list(props)[:20],
        "sample_features": features[:5],
    }


def _collect_coords(geom: dict, out: list):
    """Recursively collect coordinate pairs from a GeoJSON geometry."""
    gtype = geom.get("type", "")
    coords = geom.get("coordinates")
    if gtype == "Point" and coords:
        out.append(coords[:2])
    elif gtype in ("LineString", "MultiPoint") and coords:
        for c in coords:
            if isinstance(c, (list, tuple)) and len(c) >= 2:
                out.append(c[:2])
    elif gtype in ("Polygon", "MultiLineString") and coords:
        for ring in coords:
            for c in ring:
                if isinstance(c, (list, tuple)) and len(c) >= 2:
                    out.append(c[:2])
    elif gtype == "MultiPolygon" and coords:
        for poly in coords:
            for ring in poly:
                for c in ring:
                    if isinstance(c, (list, tuple)) and len(c) >= 2:
                        out.append(c[:2])
    elif "geometries" in geom:
        for g in geom["geometries"]:
            _collect_coords(g, out)


def _inspect_csv(path: str) -> dict:
    import csv as csv_mod
    with open(path, "r", newline="", encoding="utf-8-sig") as f:
        reader = csv_mod.reader(f)
        headers = next(reader, [])
        rows = []
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= 5000:
                break

    # Basic stats
    meta = {
        "type": "csv",
        "rows": len(rows),
        "columns": headers,
    }

    # Detect lat/lon columns
    lat_cols = [c for c in headers if c.lower() in ("lat", "latitude", "y", "lat_y")]
    lon_cols = [c for c in headers if c.lower() in ("lon", "lng", "longitude", "x", "lon_x")]
    meta["has_latlon"] = bool(lat_cols and lon_cols)

    # Detect numeric columns and compute stats
    numeric_stats = {}
    for ci, col in enumerate(headers):
        vals = []
        for row in rows:
            if ci < len(row):
                try:
                    vals.append(float(row[ci]))
                except (ValueError, TypeError):
                    pass
        if vals and len(vals) > len(rows) * 0.5:
            numeric_stats[col] = {
                "min": min(vals),
                "max": max(vals),
                "mean": sum(vals) / len(vals),
                "count": len(vals),
            }
    meta["numeric_stats"] = numeric_stats

    # Sample rows (as dicts)
    sample = []
    for row in rows[:30]:
        d = {}
        for ci, col in enumerate(headers):
            d[col] = row[ci] if ci < len(row) else ""
        sample.append(d)
    meta["sample_rows"] = sample

    return meta


def _inspect_excel(path: str) -> dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        row_count = ws.max_row - 1 if ws.max_row else 0
        wb.close()
        return {
            "type": "excel",
            "rows": row_count,
            "columns": headers,
            "sheet_name": ws.title,
        }
    except Exception:
        return {"type": "excel"}


def _inspect_tif(path: str) -> dict:
    try:
        import rasterio
        with rasterio.open(path) as ds:
            import numpy as np
            # Read small sample for stats
            data = ds.read(1, out_shape=(min(ds.height, 200), min(ds.width, 200)))
            valid = data[data != ds.nodata] if ds.nodata is not None else data
            return {
                "type": "tif",
                "bands": ds.count,
                "width": ds.width,
                "height": ds.height,
                "crs": str(ds.crs),
                "bounds": list(ds.bounds),
                "dtype": str(ds.dtypes[0]),
                "nodata": ds.nodata,
                "stats": {
                    "min": float(valid.min()) if valid.size else None,
                    "max": float(valid.max()) if valid.size else None,
                    "mean": float(valid.mean()) if valid.size else None,
                },
            }
    except Exception:
        return {"type": "tif"}
